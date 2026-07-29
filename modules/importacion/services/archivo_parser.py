"""Parseo de archivos CSV / XLSX en memoria (sin persistir path)."""
import csv
import io
from typing import List, Dict, Tuple


class ArchivoInvalido(Exception):
    pass


def _decode(contenido: bytes) -> str:
    for enc in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return contenido.decode(enc)
        except UnicodeDecodeError:
            continue
    # ultimo recurso: reemplaza caracteres invalidos
    return contenido.decode("utf-8", errors="replace")


def _parse_csv(contenido: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    texto = _decode(contenido)
    # deteccion simple de delimitador
    sample = texto[:4096]
    delim = ","
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        delim = dialect.delimiter
    except Exception:
        # fallback: el que mas aparezca en la primer linea
        primera = sample.splitlines()[0] if sample.splitlines() else ""
        delim = max([",", ";", "\t", "|"], key=lambda d: primera.count(d))
    reader = csv.reader(io.StringIO(texto), delimiter=delim)
    filas = [r for r in reader if any((c or "").strip() for c in r)]
    if not filas:
        return [], []
    headers = [(h or "").strip() or f"col_{i+1}" for i, h in enumerate(filas[0])]
    registros = []
    for fila in filas[1:]:
        registros.append({headers[i]: (fila[i] if i < len(fila) else "") for i in range(len(headers))})
    return headers, registros


def _parse_xlsx(contenido: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    ws = wb.active
    filas_iter = ws.iter_rows(values_only=True)
    try:
        primera = next(filas_iter)
    except StopIteration:
        return [], []
    headers = [(str(h).strip() if h is not None else f"col_{i+1}") for i, h in enumerate(primera)]
    registros = []
    for fila in filas_iter:
        if fila is None or all(c is None or str(c).strip() == "" for c in fila):
            continue
        reg = {}
        for i, h in enumerate(headers):
            val = fila[i] if i < len(fila) else None
            reg[h] = "" if val is None else str(val)
        registros.append(reg)
    wb.close()
    return headers, registros


def parse(nombre_archivo: str, contenido: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    """Devuelve (headers, [ {header: valor} ]) segun la extension del archivo."""
    nombre = (nombre_archivo or "").lower()
    if nombre.endswith(".xlsx") or nombre.endswith(".xlsm"):
        return _parse_xlsx(contenido)
    if nombre.endswith(".csv") or nombre.endswith(".txt"):
        return _parse_csv(contenido)
    # sin extension reconocida: intento CSV, si falla XLSX
    try:
        return _parse_csv(contenido)
    except Exception:
        return _parse_xlsx(contenido)
